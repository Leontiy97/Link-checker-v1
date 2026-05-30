import asyncio
from asyncio import Semaphore
import time


import openpyxl

from parser.link_cheker_v1 import LinkChecker



async def first_script(ref_page, page_link, anchor_text):
    httpx_result = LinkChecker(ref_page, page_link, anchor_text)
    result = await httpx_result.run()
    return result

async def process_single_row(sem, ref_page, page_link, anchor, row_index):
    async with sem:
        try:
            result = await first_script(ref_page, page_link, anchor)
            print(result)
            return row_index, result
        except Exception as e:
            print(f"Error in row {row_index}: {e}")
            from parser.verdicts import Verdicts
            return row_index, Verdicts.SYSTEM_ERROR

async def run_from_excel(filepath):
    start = time.perf_counter()
    sem = Semaphore(4)
    excel = openpyxl.load_workbook(filepath)
    excel_sheet = excel.active
    headers = {cell.value: index for index, cell in enumerate(excel_sheet[1], start=1)}
    col_ref = headers["Ref page"]
    col_anchor = headers["Anchor"]
    col_link = headers["Project url"]
    col_verdict = headers["Verdict"]
    tasks = []

    for row_idx, row in enumerate(excel_sheet.iter_rows(min_row=2, values_only=False), start=2):
        ref_page = row[col_ref - 1].value
        anchor = row[col_anchor - 1].value
        page_link = row[col_link - 1].value
        row_index = row_idx

        task = process_single_row(sem, ref_page, page_link, anchor, row_index)
        tasks.append(task)
    results = await asyncio.gather(*tasks)
    for row_index, verdict in results:
        excel_sheet.cell(row=row_index, column=col_verdict).value = verdict.value
    excel.save(filepath)
    end = time.perf_counter()
    total_seconds = end - start
    minutes, seconds = divmod(total_seconds, 60)
    print(f"Total run time: {int(minutes)} min {seconds:.2f} sec")
